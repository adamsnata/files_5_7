from openpyxl import load_workbook
import os.path
from conftest import xlsx_file, RESOURCE_PATH
# TODO оформить в тест, добавить ассерты и использовать универсальный путь

def test_xlsx():
    if os.path.exists(RESOURCE_PATH):
        if os.path.isfile(xlsx_file):
            workbook = load_workbook(xlsx_file)
            sheet = workbook.active
            print(sheet.cell(row=3, column=2).value)
        else:
            print("Файл не найден.")

        assert sheet.cell(row=3, column=2).value == 'Mara'